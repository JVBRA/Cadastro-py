from logging import PlaceHolder
import customtkinter as ctk
import tkinter as tk
from tkinter import messagebox

list_email=[]#onde ficara armazenada os dados
list_senha=[]#onde ficara armazenada os dados
email_entry=[]
senha_entry=[]

janela = ctk.CTk()
janela.geometry("325x500")

def criar_user():
    from criar import janela_criar
    # Criar uma instância da janela de criação de usuário
    janela_cri = janela_criar()
    janela_cri.executar()

    
def clique():
    #--------------------------------------------------
    import openpyxl as xl 
    book = xl.load_workbook('planilha_user.xlsx')
    user_page = book['usernames']
    senha_entry=senha.get()
    email_entry=email.get()
    print(senha_entry)
    print(email_entry)
    for column in user_page.iter_cols(min_row=2,max_row=500):
        for cell in column:
            if cell.value == senha_entry:
                coluna = True


    for row in user_page.iter_rows(min_row=2,max_row=500):
        for cell in row:
            if cell.value == str(email_entry) and coluna == True:
                from Project import janela_cadastro
                janela_cad = janela_cadastro()
                janela_cad.executar()  
    else: 
        messagebox.showinfo("Erro","E-mail ou senha Errada!")

texto = ctk.CTkLabel(janela, text="Fazer Login")
texto.grid(row=3, column=0,padx = 10, pady=10, sticky='nswe', columnspan =4 )

email= ctk.CTkEntry(janela, placeholder_text="Seu Email")
email.grid(row=4, column=0,padx = 10, pady=10, sticky='nswe', columnspan =4 )


senha= ctk.CTkEntry(janela, placeholder_text="Sua Senha",show="*")
senha.grid(row=5, column=0,padx = 10, pady=10, sticky='nswe', columnspan =4 )

botao = ctk.CTkButton(janela, text="Login", command=clique)
botao.grid(row=6, column=0,padx = 10, pady=10, sticky='nswe', columnspan =2 )


botao_new = ctk.CTkButton(janela, text="Criar Cadastro", command=criar_user)
botao_new.grid(row=6, column=2,padx = 10, pady=10, sticky='nswe', columnspan =2 )

janela.mainloop()